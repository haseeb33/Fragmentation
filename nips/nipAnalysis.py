#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Updated on Sat Jan 20 17:35:22 2020

@author: khan
"""

import pandas as pd
import matplotlib.pyplot as plt
import re
from scipy.signal import savgol_filter
from openpyxl import load_workbook

file_name = "60topics.xlsx"
wb = load_workbook(file_name, read_only=True)
TOPICS = int(file_name[:2])
TIME_SLICES = 31
T = [i for i in range(TIME_SLICES)] # Total time slices
dtm_gl = 0 # global value: changed with graph() function
year_col = 60 # year's column in sheet name "Theta"
d_years = {} # count of docs in each year, filled with function Init_d_years() call

def removeSpecialChar(orignal):
    return re.sub(r"[^a-zA-Z0-9.]+", '', orignal)
def strToval(a):
    w,m = a.split(",")
    return removeSpecialChar(w)
def J(var): # use in graphJS() function
    return JS[var][dtm_gl*TIME_SLICES:(dtm_gl+1)*TIME_SLICES]
def DTMWithAllLDA(a): # use in DTMcount(), can also be called indiviually
    return [i[1] for i in DTM_related_LDA if i[0] == a]

def populationGraph(t, mdoel="LDA"): # graph of estimated number of docs for topic t of LDA or DTM
    d = {}
    distribution = Theta_ls
    if model == "DTM":
        distribution = Gemma_ls
    for p in distribution:
        if p[year_col] not in d.keys():
            d[p[year_col]] = p[t]
            d_years[p[year_col]] = 1
        else:
            d[p[year_col]] += p[t]
            d_years[p[year_col]] += 1

    sorted_d_keys = sorted(d.keys())
    output = [d[i]/d_years[i] for i in sorted_d_keys]
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.plot(output, label = "Original")
    y = savgol_filter(output, 15, 3)
    ax.plot(y, label = "Smooth")
    ax.legend()
    plt.show()

def Init_d_years(): #fill d_years dict
    for p in Theta_ls:
        if p[year_col] not in d_years.keys():
            d_years[p[year_col]] = 1

def timeCorr(t1,t2):
    corr_time = []

    for i in sorted(d_years.keys()):
        a = Theta_df[Theta_df[TOPICS]==i]
        corr_time.append(a[t1].corr(a[t2]))
    fig = plt.figure()
    ax = fig.add_subplot(111)
    ax.set_ylim(-0.75,0.75)
    ax.plot(corr_time)

def highThetaCorr(t1, t2):
    corr_time = []
    for i in sorted(d_years.keys()):
        a = Theta_df[Theta_df[TOPICS]==i]
        corr_time.append(a[a[t1]>.015][t1].corr(a[a[t2]>.015][t2]))
    plt.plot(corr_time)

def graphJS(dtm, ls): #ls is list of LDA topics,
    global dtm_gl
    dtm_gl = dtm
    fig = plt.figure()
    ax = fig.add_subplot(111)
    plt.ylabel("JSD value")
    plt.xlabel("Time slot")
    for i in ls:
        ax.plot(T, J(i), label="Topic %i" %i)
    ax.legend()
    plt.show()

def getAllDTMwords(topic):
    words = []
    for i in range(TIME_SLICES):
        for j in range(50):
            w = strToval(DTM[i*TOPICS+topic][j])
            if w not in words:
                words.append(w)
    return words

def DTMcount():
    DTM_count_dict = {}

    for i in DTM_related_LDA:
        if i[0] in DTM_count_dict.keys():
            DTM_count_dict[i[0]] += 1
        else:
            DTM_count_dict[i[0]] = 1
    DTM_2orMore = [[dtm, DTMWithAllLDA(dtm)] for dtm, v in DTM_count_dict.items() if v>1]
    return DTM_count_dict, DTM_2orMore

def getDTM(topic):
    print([strToval(DTM[topic][i]) for i in range(50)])
def getLDA(topic):
    print([strToval(LDA[topic][i]) for i in range(50)])

def comp(l1, l2):
    print("Words removed :----->")
    for i in l1:
        if i not in l2:
            print(i , end = ", ")
    print("\n" + "Words added :------>")
    for j in l2:
        if j not in l1:
            print(j, end = ", ")
def common(l1, l2):
    print("Common words in both: ---->")
    for i in l1:
        if i in l2:
            print(i, end = ", ")

f = pd.ExcelFile(file_name)
DTM = pd.read_excel(f, sheet_name = "DTM", index_col=0).values.tolist()
LDA = pd.read_excel(f, sheet_name = "LDA", index_col=0).values.tolist()
DTM_time = pd.read_excel(f, sheet_name = "DTMTime").values.tolist()[0][1]
LDA_time = pd.read_excel(f, sheet_name = "LDATime").values.tolist()[0][1]
JS = pd.read_excel(f, sheet_name = "JS", index_col=0).values.tolist()
if 'Theta' in wb.sheetnames:
    Theta_df = pd.read_excel(f, sheet_name = "Theta", index_col=0)
    Theta_ls = Theta_df.values.tolist()
if 'DTMDistribution' in wb.sheetnames:
    Gemma_df = pd.read_excel(f, sheet_name = "DTMDistribution", index_col=0)
    Gemma_ls = Gemma_df.values.tolist()

DTM_related_LDA = []

for i_index, i in enumerate(JS):
    for j_index, j in enumerate(i):
        if j<=0.7:
            a = [j_index//TIME_SLICES, i_index]
            if a not in DTM_related_LDA:
                DTM_related_LDA.append(a)
DTM_related_LDA = sorted(DTM_related_LDA)
Init_d_years() #function call to fill d_years dict
DTM_all, DTM_2orMore = DTMcount()
